import pandas as pd
import os
import re
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pytesseract
import cv2
from dotenv import load_dotenv
import shutil
import datetime
import json
try:
    from openai import OpenAI
except Exception:
    OpenAI = None

# Load environment variables from a .env file located next to this script (if present).
# Validate the .env contents first to give a clearer message if parsing fails.
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    try:
        with open(_env_path, "r", encoding="utf-8") as _f:
            _lines = _f.readlines()
    except Exception:
        # fallback to system default encoding
        with open(_env_path, "r", encoding="utf-8", errors="replace") as _f:
            _lines = _f.readlines()

    _bad = []
    import re as _re
    for _i, _raw in enumerate(_lines, start=1):
        _s = _raw.strip()
        if not _s or _s.startswith("#"):
            continue
        # handle BOM on first line
        if _i == 1 and _s.startswith("\ufeff"):
            _s = _s.lstrip("\ufeff")
        # valid pattern starts with KEY=
        if not _re.match(r'^[A-Za-z_][A-Za-z0-9_]*\s*=.*$', _s):
            _bad.append((_i, _raw.rstrip("\n")))

    if _bad:
        print("⚠ .env parsing issues: the following lines don't match KEY=VALUE format:")
        for _ln, _content in _bad:
            print(f"  Line {_ln}: {_content!r}")
        print("Please fix the .env file (remove shell prefixes like 'export' or 'set', ensure KEY=VALUE),")
        print("or save the file as UTF-8 without BOM. Skipping dotenv.load_dotenv() to avoid noisy warnings.")
    else:
        load_dotenv(_env_path)
else:
    # no .env present — nothing to load
    pass

# ---------------- CONFIG ----------------
# Configure pytesseract executable path dynamically:
# 1. TESSERACT_CMD env (set by Electron with bundled path, or user override)
# 2. TESSERACT_PATH env
# 3. NCRP_APP_PATH + relative path (Electron bundled: resources/tools/tesseract/tesseract.exe)
# 4. shutil.which('tesseract')
# 5. Common Windows install paths
def _resolve_tesseract_path():
    tess_env = os.environ.get('TESSERACT_CMD') or os.environ.get('TESSERACT_PATH')
    if tess_env and os.path.exists(tess_env):
        return tess_env
    # Bundled path (Electron sets NCRP_APP_PATH to resources folder)
    app_path = os.environ.get('NCRP_APP_PATH')
    if app_path:
        for rel in ['tools/tesseract/tesseract.exe', 'tesseract/tesseract.exe']:
            candidate = os.path.join(app_path, rel)
            if os.path.exists(candidate):
                return os.path.normpath(candidate)
    which_tess = shutil.which('tesseract')
    if which_tess:
        return which_tess
    # Common Windows paths (no user-specific paths)
    for _path in [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.join(os.path.expandvars("%ProgramFiles%"), "Tesseract-OCR", "tesseract.exe"),
        os.path.join(os.path.expandvars("%ProgramFiles(x86)%"), "Tesseract-OCR", "tesseract.exe"),
        os.path.join(os.path.expandvars("%LOCALAPPDATA%"), "Programs", "Tesseract-OCR", "tesseract.exe"),
    ]:
        if _path and os.path.exists(_path):
            return _path
    return None

_tess_cmd = _resolve_tesseract_path()

if _tess_cmd:
    pytesseract.pytesseract.tesseract_cmd = _tess_cmd

    tess_dir = os.path.dirname(os.path.abspath(_tess_cmd))
    tessdata_dir = os.path.join(tess_dir, "tessdata")

    if not os.path.exists(os.path.join(tessdata_dir, "eng.traineddata")):
        raise RuntimeError(f"eng.traineddata not found in {tessdata_dir}")

    # TESSDATA_PREFIX = tessdata folder (Tesseract looks for TESSDATA_PREFIX/eng.traineddata)
    os.environ["TESSDATA_PREFIX"] = os.path.normpath(tessdata_dir).replace("\\", "/")

else:
    raise RuntimeError("Tesseract executable not found")

# Excel output: C:\NCRP\ncrp_complaints.xlsx (or NCRP_DATA_PATH when set by Electron)
_NCRP_BASE = os.environ.get('NCRP_DATA_PATH', r'C:\NCRP')
os.makedirs(_NCRP_BASE, exist_ok=True)
OUTPUT_FILE = os.path.join(_NCRP_BASE, 'ncrp_complaints.xlsx')

COLUMNS = [
    "Source",
    "Complaint ID",
    "Complaint Date",
    "Incident Date & Time",
    "Mobile",
    "Email",
    "Full Address",
    "District",
    "State",
    "Cybercrime Type",
    "Platform",
    "Total Amount Lost",
    "Current Status"
]

# OpenAI client — set OPENAI_API_KEY in environment to use AI parsing
try:
    client = OpenAI()
except Exception:
    client = None


# ---------------- AI PARSER (optional)
def ai_parse_complaint(text, source="IMAGE"):
    prompt = f"""
You are given raw OCR text from an Indian NCRP complaint form.
Your task is to extract ALL fields accurately.

Fields:
- Complaint ID
- Complaint Date (dd/mm/yyyy)
- Incident Date & Time
- Mobile (first valid 10-digit)
- Email
- Full Address
- District
- State
- Cybercrime Type
- Platform (Bank/UPI/Other)
- Total Amount Lost
- Current Status (Registered/Under Process)

Return strictly in JSON:
{{
"Source": "{source}",
"Complaint ID": "...",
"Complaint Date": "...",
"Incident Date & Time": "...",
"Mobile": "...",
"Email": "...",
"Full Address": "...",
"District": "...",
"State": "...",
"Cybercrime Type": "...",
"Platform": "...",
"Total Amount Lost": "...",
"Current Status": "..."
}}

Raw text:
{text}
"""
    if client is None:
        print("⚠ OpenAI client not available (OPENAI_API_KEY missing). Skipping AI parse.")
        return {col: "NOT FOUND" for col in COLUMNS}
    try:
        res = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        content = res.choices[0].message.content
        m = re.search(r"\{.*\}", content, re.DOTALL)
        if m:
            data = json.loads(m.group())
            return {k: v if v else "NOT FOUND" for k, v in data.items()}
        else:
            return {col: "NOT FOUND" for col in COLUMNS}
    except Exception as e:
        print(f"⚠ AI parsing failed: {e}")
        return {col: "NOT FOUND" for col in COLUMNS}

# ---------------- HELPERS ----------------
def clean(text):
    return re.sub(r"\s+", " ", text).strip()

def safe(val):
    return val if val else "NOT FOUND"

def first_match(patterns, text):
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return clean(m.group(1) if m.lastindex else m.group(0))
    return ""

# ---------------- READERS ----------------
def read_pdf(path):
    reader = PdfReader(path)
    text = ""
    for page in reader.pages:
        t = page.extract_text()
        if t:
            text += " " + t
    return clean(text)

def _get_tessdata_config():
    """Build config for --tessdata-dir (points directly to tessdata)."""
    if _tess_cmd and os.path.exists(_tess_cmd):
        tess_dir = os.path.dirname(os.path.abspath(_tess_cmd))
        tessdata_dir = os.path.join(tess_dir, "tessdata")

        if (
            os.path.isdir(tessdata_dir)
            and os.path.exists(os.path.join(tessdata_dir, "eng.traineddata"))
        ):
            return f'--tessdata-dir "{tessdata_dir.replace(chr(92), "/")}"'

    return ""


def read_image(path):
    img = cv2.imread(path)
    if img is None:
        raise ValueError(f"Image not readable: {path}")

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)[1]

    # DO NOT pass --tessdata-dir
    return clean(
        pytesseract.image_to_string(
            gray,
            lang="eng",
            config="--oem 3 --psm 6"
        )
    )

# ---------------- EXTRACTION ----------------
def extract_ncrp(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    text = read_pdf(file_path) if ext == ".pdf" else read_image(file_path)
    source = "PDF" if ext == ".pdf" else "IMAGE"

    complaint_id = safe(first_match([
        r"Acknowledgement Number\s*[:\-]?\s*(\d+)",
        r"Complaint ID\s*[:\-]?\s*(\d+)",
        r"\b\d{10,}\b"
    ], text))

    complaint_date = safe(first_match([
        r"Complaint Date\s*[:\-]?\s*([0-9 ]{1,2}/[0-9 ]{1,2}/[0-9]{4})"
    ], text))

    incident_dt = safe(first_match([
        r"Incident Date\/Time\s*[:\-]?\s*([0-9 ]{1,2}/[0-9 ]{1,2}/[0-9]{4}\s+[0-9 ]{1,2}\s*:\s*[0-9 ]{1,2}\s*:\s*[0-9 ]{1,2}\s*[APMapm]{2})",
        r"Incident Date\s*[:\-]?\s*([0-9 ]{1,2}/[0-9 ]{1,2}/[0-9]{4})"
    ], text))

    mobile = safe(first_match([
        r"Mobile\s*[:\-]?\s*(\d{9,10})",
        r"\b\d{9,10}\b"
    ], text))

    email = safe(first_match([
        r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    ], text))

    district = safe(first_match([
        r"District\s*[:\-]?\s*([A-Za-z ]+)"
    ], text))

    state = safe(first_match([
        r"State\s*[:\-]?\s*([A-Za-z ]+)"
    ], text))

    total_amt = safe(first_match([
        r"Total Fraudulent Amount.*?:\s*([\d,\.]+)",
        r"Total Amount.*?:\s*([\d,\.]+)"
    ], text))

    cat = first_match([r"Category of complaint\s*(.+?)\sSub"], text)
    sub = first_match([r"Sub Category of Complaint\s*(.+?)\s"], text)

    address_parts = [
        first_match([r"House No\s*[:\-]?\s*(.+?)\s"], text),
        first_match([r"Street Name\s*[:\-]?\s*(.+?)\s"], text),
        first_match([r"Village\/Town\s*[:\-]?\s*(.+?)\s"], text),
        first_match([r"Pincode\s*[:\-]?\s*(\d+)"], text)
    ]
    address = safe(", ".join(dict.fromkeys(filter(None, address_parts))))

    platform = "UPI" if "UPI" in text.upper() else "Bank" if "BANK" in text.upper() else "Other"
    status = "Under Process" if "UNDER PROCESS" in text.upper() else "Registered"

    return {
        "Source": source,
        "Complaint ID": complaint_id,
        "Complaint Date": complaint_date,
        "Incident Date & Time": incident_dt,
        "Mobile": mobile,
        "Email": email,
        "Full Address": address,
        "District": district,
        "State": state,
        "Cybercrime Type": safe(f"{cat} - {sub}".strip(" -")),
        "Platform": platform,
        "Total Amount Lost": total_amt,
        "Current Status": status
    }

# ---------------- MAIN ----------------


def save_df_to_mysql(df, *, user, password, host="localhost", port=3306, db="ncrp_db", table="ncrp_complaints"):
    """
    Save pandas DataFrame to MySQL using SQLAlchemy + pymysql.
    - Renames DataFrame columns to snake_case matching typical SQL column names.
    - Cleans and converts date/amount columns.
    - Appends rows to the target table.
    """
    # Local imports so the main script can run without these packages if DB saving is not used
    from sqlalchemy import create_engine
    from sqlalchemy.types import String, Date, DateTime, Text, DECIMAL
    import pandas as pd

    # 1) Rename columns from the DataFrame-friendly names to SQL-friendly names
    col_map = {
        "Source": "source",
        "Complaint ID": "complaint_id",
        "Complaint Date": "complaint_date",
        "Incident Date & Time": "incident_datetime",
        "Mobile": "mobile",
        "Email": "email",
        "Full Address": "full_address",
        "District": "district",
        "State": "state",
        "Cybercrime Type": "cybercrime_type",
        "Platform": "platform",
        "Total Amount Lost": "total_amount_lost",
        "Current Status": "current_status",
    }

    df_db = df.rename(columns=col_map).copy()

    # 2) Clean / convert columns
    if "complaint_date" in df_db.columns:
        df_db["complaint_date"] = pd.to_datetime(df_db["complaint_date"], dayfirst=True, errors="coerce").dt.date

    if "incident_datetime" in df_db.columns:
        df_db["incident_datetime"] = pd.to_datetime(df_db["incident_datetime"], dayfirst=True, errors="coerce")

    if "total_amount_lost" in df_db.columns:
        df_db["total_amount_lost"] = (
            df_db["total_amount_lost"].astype(str)
            .str.replace(",", "", regex=False)
            .str.replace(" ", "", regex=False)
            .replace({"NOT FOUND": None})
        )
        df_db["total_amount_lost"] = pd.to_numeric(df_db["total_amount_lost"], errors="coerce")

    # 3) Build SQLAlchemy engine
    url = f"mysql+pymysql://{user}:{password}@{host}:{port}/{db}?charset=utf8mb4"
    engine = create_engine(url, pool_recycle=3600)

    # 4) Provide dtype mapping for to_sql
    dtype_map = {
        "source": String(32),
        "complaint_id": String(128),
        "complaint_date": Date(),
        "incident_datetime": DateTime(),
        "mobile": String(32),
        "email": String(256),
        "full_address": Text(),
        "district": String(128),
        "state": String(128),
        "cybercrime_type": String(256),
        "platform": String(64),
        "total_amount_lost": DECIMAL(18, 2),
        "current_status": String(64),
    }

    # 5) Write to DB
    # NOTE: if the table does not exist, you can run a CREATE TABLE manually or use if_exists='replace' once.
    df_db.to_sql(name=table, con=engine, if_exists='append', index=False, dtype=dtype_map, method='multi', chunksize=500)

    engine.dispose()

if __name__ == "__main__":
    files = [f for f in os.listdir() if f.lower().endswith((".pdf", ".jpg", ".jpeg", ".png"))]

    if not files:
        print("❌ No PDF or Image files found in folder")
        exit()

    # Optional: use AI parsing pipeline (from ncrp_automation) when USE_AI env var is set
    if os.environ.get("USE_AI", "").lower() in ("1", "true", "yes"):
        pdfs = [f for f in files if f.lower().endswith(".pdf")]
        images = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png"))]

        all_rows = []

        # Process each PDF individually with AI
        for pdf in pdfs:
            try:
                print(f"🔍 AI Processing PDF: {pdf}")
                text = read_pdf(pdf)
                all_rows.append(ai_parse_complaint(text, source="PDF"))
            except Exception as e:
                print(f"⚠ Failed AI parse on {pdf}: {e}")

        # Merge images (WhatsApp) into a single complaint and parse with AI
        if images:
            try:
                print("🔍 AI Processing WhatsApp images (merged)")
                merged_text = ""
                for img in sorted(images):
                    merged_text += " " + read_image(img)
                all_rows.append(ai_parse_complaint(merged_text, source="IMAGE"))
            except Exception as e:
                print(f"⚠ Failed AI parse on images: {e}")

        if not all_rows:
            print("❌ No complaints extracted by AI")
            exit()

        # Save all complaints to a single Excel and format
        all_excel = "ncrp_complaints_ALL.xlsx"
        df_all = pd.DataFrame(all_rows, columns=COLUMNS)
        df_all.to_excel(all_excel, index=False)
        try:
            wb = load_workbook(all_excel)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
            wb.save(all_excel)
        except Exception as e:
            print(f"⚠ Failed formatting all Excel file: {e}")
        print(f"✔ All complaints saved in {all_excel}")

        # Save per-platform Excel files
        platforms = set([row.get("Platform", "Other") for row in all_rows])
        for plat in platforms:
            rows_plat = [r for r in all_rows if r.get("Platform", "Other") == plat]
            if not rows_plat:
                continue
            dfp = pd.DataFrame(rows_plat, columns=COLUMNS)
            output_file = f"ncrp_complaints_{plat}.xlsx"
            dfp.to_excel(output_file, index=False)
            try:
                wb = load_workbook(output_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
                wb.save(output_file)
            except Exception as e:
                print(f"⚠ Failed formatting {output_file}: {e}")
            print(f"✔ {len(rows_plat)} complaints saved to {output_file}")

        print(f"✔ SUCCESS — Total {len(all_rows)} complaints processed by AI")
        exit()

    rows = []
    for f in files:
        try:
            print(f"🔍 Processing: {f}")
            rows.append(extract_ncrp(f))
        except Exception as e:
            print(f"⚠ Failed on {f}: {e}")

    df = pd.DataFrame(rows, columns=COLUMNS)

    # Helper: save Excel safely. If the target file is locked (PermissionError),
    # write to a timestamped fallback file and return that path.
    def _safe_save_excel(df_obj, target_path):
        try:
            out_dir = os.path.dirname(target_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            df_obj.to_excel(target_path, index=False)
            return target_path
        except PermissionError:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(target_path)
            alt = f"{base}_{ts}{ext}"
            try:
                df_obj.to_excel(alt, index=False)
                print(f"⚠ Permission denied writing '{target_path}'. Saved to fallback file: '{alt}'")
                return alt
            except Exception as e:
                # re-raise the original PermissionError context for visibility
                raise PermissionError(f"Failed to write both primary ({target_path}) and fallback ({alt}): {e}") from e

    out_path = _safe_save_excel(df, OUTPUT_FILE)

    # Post-process formatting: try to open the file, but skip if it's locked.
    try:
        wb = load_workbook(out_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        try:
            wb.save(out_path)
        except PermissionError:
            print(f"⚠ Permission denied when saving formatting to '{out_path}'. The file may be open in Excel. Formatting skipped.")
    except PermissionError:
        print(f"⚠ Permission denied when opening '{out_path}' for formatting. The file may be open in Excel. Skipping formatting.")
    except Exception as e:
        print(f"⚠ Failed during Excel post-processing: {e}")

    print("✔ ALL DATA EXTRACTED — Complaint Date & Incident Date/Time fixed")

    # --- Optional: attempt to save to MySQL if environment variables are set ---
    # To enable automatic DB save, set the following environment variables before running:
    # DB_USER, DB_PASSWORD, DB_NAME (optional), DB_HOST (optional), DB_PORT (optional), DB_TABLE (optional)
    db_user = os.environ.get("DB_USER")
    # allow empty password (DB_PASSWORD="" in .env) — default to empty string
    db_password = os.environ.get("DB_PASSWORD", "")
    # proceed if DB_USER is provided (even if password is empty)
    if db_user:
        db_host = os.environ.get("DB_HOST", "localhost")
        db_port = int(os.environ.get("DB_PORT", 3306))
        db_name = os.environ.get("DB_NAME", "ncrp_db")
        db_table = os.environ.get("DB_TABLE", "ncrp_complaints")

        try:
            print(f"🔁 Saving extracted data to MySQL... (host={db_host}, db={db_name}, table={db_table}, user={db_user})")
            save_df_to_mysql(df,
                             user=db_user,
                             password=db_password,
                             host=db_host,
                             port=db_port,
                             db=db_name,
                             table=db_table)
            print(f"✔ Data saved to MySQL — rows attempted: {len(df)}")
        except Exception as e:
            print(f"⚠ Failed to save to MySQL: {e}")
    else:
        print("ℹ To save to MySQL automatically, set DB_USER in your environment or in .env (DB_PASSWORD may be empty)")
